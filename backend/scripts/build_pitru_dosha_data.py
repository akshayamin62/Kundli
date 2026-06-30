"""Regenerate pitru_dosha_data.py from Excel (+ severity from House Matrix & Dosha Matrix)."""
import os
import re
import sys

import openpyxl

SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, SCRIPTS_DIR)

from pitru_dosha_remedies import enrich_domain, remedies_for_domain  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX_MAIN = os.path.join(ROOT, "Pitru_Dosha_Zodiac_House_Wise_Combinations.xlsx")
XLSX_DOSHA = os.path.join(ROOT, "Pitru_Dosha_Health_and_Modern_Solutions.xlsx")
XLSX_CAREER = os.path.join(ROOT, "Pitrudosha and Career.xlsx")
XLSX_FINANCE = os.path.join(ROOT, "Pitrudosha and Finance.xlsx")
XLSX_RELATIONSHIP = os.path.join(ROOT, "Pitrudosha and Relationship.xlsx")
OUT = os.path.join(ROOT, "backend", "app", "data", "pitru_dosha_data.py")

DOMAIN_KEYS = ("health", "career", "finance", "relationship")

# Sign Wise combination → Dosha Matrix severity (same for every zodiac sign row).
SIGN_COMBINATION_SEVERITY: dict[str, str] = {
    "Sun + Rahu": "High",
    "Sun + Ketu": "Medium-High",
    "Saturn + Rahu": "Very High",
    "Saturn + Ketu": "High",
    "Moon + Rahu": "Medium-High",
    "Moon + Ketu": "Medium-High",
    "Debilitated Sun + Rahu": "High",
    "Rahu/Ketu affecting 9th": "High",
    "Rahu/Ketu in Taurus-Scorpio axis": "Very High",
    "Rahu/Ketu in Gemini-Sagittarius axis": "High",
    "Rahu/Ketu on Cancer-Capricorn axis": "Medium",
    "Rahu/Ketu in Cancer-Capricorn axis": "Medium",
    "Rahu/Ketu in Libra-Aries axis": "Medium-High",
    "Rahu/Ketu in Virgo-Pisces axis": "High",
    "Mercury afflicted with Rahu/Ketu": "Medium-High",
    "Mercury + Rahu/Ketu": "High",
    "Afflicted Venus with Rahu/Ketu": "Medium-High",
    "Afflicted Venus": "Medium-High",
    "Afflicted Mars": "Medium-High",
    "Afflicted Mars linked to 8th/9th": "Very High",
    "Afflicted Jupiter": "Medium-High",
    "9th lord in Virgo dusthana": "High",
    "Afflicted 9th lord in Capricorn": "High",
    "Saturn aspect on Sun": "High",
    "Saturn influence on Cancer Moon": "Medium-High",
    "Rahu/Ketu affecting Leo 5th": "High",
    "Sun weak in D-9/D-12 despite Leo placement": "Medium-High",
}

# House Wise combination → Dosha Matrix severity (when listed); else House Severity Matrix.
HOUSE_COMBINATION_SEVERITY: dict[str, str] = {
    "Sun + Rahu in 1st": "High",
    "Sun + Rahu in 3rd": "High",
    "Sun + Rahu in 5th": "High",
    "Sun + Rahu in 7th": "High",
    "Sun + Rahu in 8th": "High",
    "Sun + Rahu in 9th": "High",
    "Sun + Rahu in 10th": "High",
    "Sun + Rahu in 12th": "High",
    "Sun + Ketu in 1st": "Medium-High",
    "Sun + Ketu in 8th": "High",
    "Sun + Ketu in 12th": "Medium-High",
    "Sun + Saturn in 9th": "High",
    "Sun + Saturn in 10th": "High",
    "Rahu in 9th": "High",
    "Ketu in 9th": "Medium-High",
    "Rahu in 8th": "Very High",
    "Rahu in 2nd": "Medium-High",
    "Ketu in 2nd": "Medium-High",
    "Saturn in 2nd": "Medium-High",
    "Saturn in 4th": "Medium",
    "Saturn in 5th": "High",
    "9th lord in 6th": "High",
    "9th lord in 8th afflicted": "High",
    "9th lord in 12th afflicted": "High",
    "Afflicted 9th lord in 1st": "High",
    "9th lord afflicted": "High",
    "Gulika/Mandi in 9th": "High",
    "Rahu/Ketu on 1st-7th axis": "Medium-High",
    "Rahu/Ketu in 1st-7th axis": "Medium-High",
    "Rahu/Ketu in 3rd-9th axis": "High",
    "Rahu/Ketu in 4th-10th axis": "Medium",
    "Rahu/Ketu in 5th-11th axis": "Medium-High",
    "Saturn + Rahu in 1st": "Very High",
    "Saturn + Rahu in 3rd": "Very High",
    "Saturn + Rahu in 6th": "Very High",
    "Saturn + Rahu in 8th": "Very High",
    "Saturn + Rahu in 10th": "High",
    "Saturn + Ketu in 7th": "High",
    "Saturn + Ketu in 12th": "High",
    "Moon + Rahu in 4th": "Medium-High",
    "Moon + Ketu in 4th": "Medium-High",
    "Sun afflicted in 4th": "Medium",
    "Sun afflicted in 6th": "High",
    "2nd lord in 8th afflicted": "Very High",
    "5th lord in 8th afflicted": "High",
    "5th lord in 8th/12th afflicted": "High",
}

# Sign/house combination → Dosha Matrix type (for remedies lookup).
COMBINATION_TO_DOSHA_TYPE: dict[str, str] = {
    "Sun + Rahu": "Sun-Rahu Pitru Dosha",
    "Sun + Ketu": "Sun-Ketu Pitru Dosha",
    "Sun + Saturn": "Sun-Saturn Pitru Dosha",
    "Moon + Rahu": "Moon-Rahu Ancestral Mental Dosha",
    "Moon + Ketu": "Moon-Ketu Emotional Detachment",
    "Saturn + Rahu": "Saturn-Rahu Ancestral Curse Type",
    "Saturn + Ketu": "Saturn-Ketu Ancestral Separation",
    "Debilitated Sun + Rahu": "Sun-Rahu Pitru Dosha",
    "Rahu/Ketu affecting 9th": "Rahu in 9th House",
    "Rahu in 9th": "Rahu in 9th House",
    "Ketu in 9th": "Ketu in 9th House",
    "Rahu in 8th": "Afflicted 8th House Lineage Dosha",
    "Rahu in 2nd": "Pitru Dosha through 2nd House",
    "Ketu in 2nd": "Pitru Dosha through 2nd House",
    "Saturn in 2nd": "Pitru Dosha through 2nd House",
    "Saturn in 4th": "Pitru Dosha through 4th House",
    "Saturn in 5th": "Afflicted 5th House",
    "9th lord in 6th": "Afflicted 9th Lord",
    "9th lord in 8th afflicted": "Afflicted 9th Lord",
    "9th lord in 12th afflicted": "Afflicted 9th Lord",
    "Afflicted 9th lord in 1st": "Afflicted 9th Lord",
    "9th lord afflicted": "Afflicted 9th Lord",
    "Afflicted 9th lord in Capricorn": "Afflicted 9th Lord",
    "9th lord in Virgo dusthana": "Afflicted 9th Lord",
    "Gulika/Mandi in 9th": "Gulika/Mandi Pitru Dosha",
    "Rahu/Ketu in Taurus-Scorpio axis": "Rahu/Ketu Axis in 2nd-8th",
    "Rahu/Ketu in Gemini-Sagittarius axis": "Rahu/Ketu Axis in 3rd-9th",
    "Rahu/Ketu on Cancer-Capricorn axis": "Rahu/Ketu Axis in 4th-10th",
    "Rahu/Ketu in Cancer-Capricorn axis": "Rahu/Ketu Axis in 4th-10th",
    "Rahu/Ketu in Libra-Aries axis": "Rahu/Ketu Axis in 1st-7th",
    "Rahu/Ketu in Virgo-Pisces axis": "Rahu/Ketu Axis in 6th-12th",
    "Rahu/Ketu on 1st-7th axis": "Rahu/Ketu Axis in 1st-7th",
    "Rahu/Ketu in 1st-7th axis": "Rahu/Ketu Axis in 1st-7th",
    "Rahu/Ketu in 3rd-9th axis": "Rahu/Ketu Axis in 3rd-9th",
    "Rahu/Ketu in 4th-10th axis": "Rahu/Ketu Axis in 4th-10th",
    "Rahu/Ketu in 5th-11th axis": "Rahu/Ketu Axis in 5th-11th",
    "Mercury afflicted with Rahu/Ketu": "Guru Chandal Yoga Pitru Link",
    "Mercury + Rahu/Ketu": "Guru Chandal Yoga Pitru Link",
    "Afflicted Venus with Rahu/Ketu": "Jupiter Afflicted Pitru Dosha",
    "Afflicted Venus": "Jupiter Afflicted Pitru Dosha",
    "Afflicted Mars linked to 8th/9th": "Sun-Mars Pitru Dosha",
    "Afflicted Mars": "Sun-Mars Pitru Dosha",
    "Afflicted Jupiter": "Jupiter Afflicted Pitru Dosha",
    "Saturn aspect on Sun": "Sun-Saturn Pitru Dosha",
    "Saturn influence on Cancer Moon": "Moon-Rahu Ancestral Mental Dosha",
    "Rahu/Ketu affecting Leo 5th": "Afflicted 5th House",
    "Sun weak in D-9/D-12 despite Leo placement": "D-9 Navamsa Pitru Indication",
    "2nd lord in 8th afflicted": "Pitru Dosha through 2nd House",
    "5th lord in 8th afflicted": "Afflicted 5th House",
    "5th lord in 8th/12th afflicted": "Afflicted 5th House",
    "Sun afflicted in 4th": "Pitru Dosha through 4th House",
    "Sun afflicted in 6th": "6th/8th/12th Link with Sun or 9th Lord",
}

PLANET_SPIRITUAL_DEFAULT: dict[str, str] = {
    "Sun": "Surya arghya, Sunday worship, respect for father and elders",
    "Moon": "Chandra puja, Monday rituals, mother-line healing",
    "Mars": "Hanuman worship, Tuesday fasting, courage through seva",
    "Mercury": "Vishnu worship, Wednesday prayers, honest speech",
    "Jupiter": "Guru seva, Thursday puja, dharma and pitru tarpan",
    "Venus": "Friday Lakshmi puja, harmony in relationships",
    "Saturn": "Saturday Shani puja, service to elders and poor",
    "Rahu": "Rahu shanti, ethical conduct, avoid shortcuts",
    "Ketu": "Ketu pacification, meditation, service to animals",
}

CONVENTIONAL_BY_DOSHA_TYPE: dict[str, str] = {
    "Sun-Rahu Pitru Dosha": "Pitru tarpan on Amavasya, Surya arghya, Rahu shanti, Shiva puja",
    "Sun-Ketu Pitru Dosha": "Pitru tarpan, Surya namaskar, Ketu peace rituals, ancestral remembrance",
    "Rahu in 9th House": "9th house peace puja, father-line tarpan, Thursday Guru worship",
    "Ketu in 9th House": "Ketu shanti, dharma seva, respect for guru and father",
    "Saturn-Rahu Ancestral Curse Type": "Shani-Rahu shanti, Saturday charity, Mahamrityunjaya japa",
    "Saturn-Ketu Ancestral Separation": "Ancestor tarpan, Ketu-Saturn peace rituals, elder seva",
    "Moon-Rahu Ancestral Mental Dosha": "Chandra shanti, Monday fasting, emotional healing prayers",
    "Moon-Ketu Emotional Detachment": "Chandra-Ketu peace, grounding sadhana, mother-line healing",
    "Jupiter Afflicted Pitru Dosha": "Guru puja, pitru tarpan, Thursday worship, dharma correction",
    "Guru Chandal Yoga Pitru Link": "Guru-Rahu/Ketu shanti, Vishnu sahasranama, ethical speech",
    "Afflicted 9th Lord": "9th lord peace rituals, father worship, pitru tarpan",
    "Afflicted 5th House": "Santan peace puja, 5th house remedies, Guru grace",
    "Afflicted 8th House Lineage Dosha": "8th house shanti, ancestral healing, Mahamrityunjaya japa",
    "Rahu/Ketu Axis in 2nd-8th": "Family lineage tarpan, 2nd-8th axis shanti, wealth dharma rituals",
    "Rahu/Ketu Axis in 3rd-9th": "Sibling and father-line peace, 3rd-9th axis remedies",
    "Rahu/Ketu Axis in 1st-7th": "Lagna shanti, marriage harmony puja, Rahu-Ketu axis peace",
    "Rahu/Ketu Axis in 4th-10th": "Griha shanti, career dharma rituals, mother-father balance",
    "Rahu/Ketu Axis in 5th-11th": "5th house Guru puja, progeny peace rituals",
    "Rahu/Ketu Axis in 6th-12th": "Health and moksha balance rituals, service to needy",
    "Pitru Dosha through 2nd House": "Family tarpan, speech discipline, 2nd house peace",
    "Pitru Dosha through 4th House": "Mother worship, griha shanti, 4th house remedies",
    "Pitru Dosha through 10th House": "Father and career dharma rituals, Sunday Surya puja",
    "Pitru Dosha through 12th House": "Moksha-oriented pitru seva, 12th house peace",
    "Gulika/Mandi Pitru Dosha": "Gulika shanti, Saturday remedies, ancestral peace",
    "Sun-Saturn Pitru Dosha": "Surya-Shani peace, Sunday-Saturday worship, father respect",
    "Sun-Mars Pitru Dosha": "Surya-Mangal shanti, Tuesday-Sunday rituals",
    "D-9 Navamsa Pitru Indication": "Navamsa peace puja, pitru tarpan, Guru worship",
    "6th/8th/12th Link with Sun or 9th Lord": "Dusthana shanti, health and pitru combined remedies",
}

DOMAIN_SHEETS: dict[str, tuple[str, str, str, str]] = {
    "career": (
        XLSX_CAREER,
        "Career Area Affected",
        "Possible Career Impact",
        "Severity",
    ),
    "finance": (
        XLSX_FINANCE,
        "Financial Area Affected",
        "Possible Impact on Income, Revenue, Savings and Wealth",
        "Severity",
    ),
    "relationship": (
        XLSX_RELATIONSHIP,
        "Relationship Area Affected",
        "Possible Relationship Impact",
        "Severity",
    ),
}


def load_modern_solutions() -> dict[str, str]:
    if not os.path.isfile(XLSX_DOSHA):
        return {}
    wb = openpyxl.load_workbook(XLSX_DOSHA, read_only=True, data_only=True)
    ws = wb["Modern Solutions"]
    out: dict[str, str] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[1] and r[3]:
            out[str(r[1]).strip()] = str(r[3]).strip()
    wb.close()
    return out


def resolve_dosha_type(combination: str) -> str | None:
    if combination in COMBINATION_TO_DOSHA_TYPE:
        return COMBINATION_TO_DOSHA_TYPE[combination]
    m = re.match(r"^(.+?) in \d+(?:st|nd|rd|th)?(?:\s+House)?$", combination, re.I)
    if m:
        base = m.group(1).strip()
        if base in COMBINATION_TO_DOSHA_TYPE:
            return COMBINATION_TO_DOSHA_TYPE[base]
    return None


def conventional_for(combination: str) -> str:
    dtype = resolve_dosha_type(combination)
    if dtype and dtype in CONVENTIONAL_BY_DOSHA_TYPE:
        return CONVENTIONAL_BY_DOSHA_TYPE[dtype]
    parts: list[str] = []
    for planet, text in PLANET_SPIRITUAL_DEFAULT.items():
        if planet in combination:
            parts.append(text)
    if parts:
        return "; ".join(dict.fromkeys(parts))
    return "Pitru tarpan on Amavasya, respect for ancestors, Shiva puja, ethical living"


def modern_for(combination: str, modern_map: dict[str, str]) -> str:
    dtype = resolve_dosha_type(combination)
    if dtype and dtype in modern_map:
        return modern_map[dtype]
    return (
        "Annual health screening as advised, family health history documentation, "
        "stress management, therapy when needed, ethical lifestyle corrections"
    )


def read_sheet(wb, name: str, header_row: int = 2) -> list[dict]:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else "" for h in rows[header_row]]
    data = []
    for r in rows[header_row + 1 :]:
        if not r or all(c is None or str(c).strip() == "" for c in r):
            continue
        d = {
            headers[i]: (str(r[i]).strip() if r[i] is not None else "")
            for i in range(len(headers))
            if headers[i]
        }
        if any(d.values()):
            data.append(d)
    return data


def read_domain_sheet(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(
        i
        for i, r in enumerate(rows)
        if r and any(c and "Pitru Dosha Combination" in str(c) for c in r)
    )
    headers = [str(h).strip() if h else "" for h in rows[header_idx]]
    data: list[dict] = []
    for r in rows[header_idx + 1 :]:
        if not r or all(c is None or str(c).strip() == "" for c in r):
            continue
        d = {
            headers[i]: (str(r[i]).strip() if r[i] is not None else "")
            for i in range(len(headers))
            if headers[i]
        }
        combo = d.get("Pitru Dosha Combination", "")
        if combo:
            data.append(d)
    wb.close()
    return data


def load_house_severity_matrix() -> dict[str, str]:
    wb = openpyxl.load_workbook(XLSX_MAIN, read_only=True, data_only=True)
    ws = wb["Severity Matrix"]
    out: dict[str, str] = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r and r[0] and r[2]:
            out[str(r[0]).strip()] = str(r[2]).strip()
    wb.close()
    return out


def house_label_to_ordinal(house_label: str) -> str:
    m = re.match(r"^(\d+)(?:st|nd|rd|th)\s+House$", house_label.strip(), re.I)
    if m:
        n = int(m.group(1))
        if 10 <= n % 100 <= 20:
            suffix = "th"
        else:
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
        return f"{n}{suffix}"
    return ""


def ordinal(n: int) -> str:
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def house_label(n: int) -> str:
    return f"{ordinal(n)} House"


def house_row_severity(house_label_str: str, combination: str, house_matrix: dict[str, str]) -> str:
    if combination in HOUSE_COMBINATION_SEVERITY:
        return HOUSE_COMBINATION_SEVERITY[combination]
    ordinal_label = house_label_to_ordinal(house_label_str)
    return house_matrix.get(ordinal_label, "")


def canonical_combination(combo: str) -> str:
    s = re.sub(r"\s+", " ", combo.strip())
    s = re.sub(r"\bin (\d+(?:st|nd|rd|th)?)\s+[Hh]ouse\b", r"in \1", s)
    s = re.sub(
        r"^(Rahu|Ketu) in (\d+(?:st|nd|rd|th)?)\s+[Hh]ouse$",
        r"\1 in \2",
        s,
        flags=re.I,
    )
    s = s.replace("Rahu/Ketu axis in", "Rahu/Ketu in")
    s = s.replace("Rahu/Ketu on ", "Rahu/Ketu in ")
    s = re.sub(r"(\d+)(?:st|nd|rd|th) house afflicted", r"\1 afflicted", s, flags=re.I)
    s = re.sub(r" with affliction$", "", s, flags=re.I)
    s = s.replace("Venus + Rahu House", "Venus + Rahu")
    s = s.replace("Venus + Ketu House", "Venus + Ketu")
    return s


def is_axis_combination(combo: str) -> bool:
    canon = canonical_combination(combo)
    return "axis" in canon.lower() or bool(
        re.search(r"Rahu/Ketu in \d+(?:st|nd|rd|th)-\d+", canon, re.I)
    )


def parse_house_label(combination: str) -> str:
    m = re.search(r"\bin (\d+)(?:st|nd|rd|th)?\s+[Hh]ouse\b", combination, re.I)
    if m:
        return house_label(int(m.group(1)))
    m = re.search(
        r"(?:^| )(?:Rahu|Ketu) in (\d+)(?:st|nd|rd|th)\s+[Hh]ouse",
        combination,
        re.I,
    )
    if m:
        return house_label(int(m.group(1)))
    m = re.search(r"(\d+)(?:st|nd|rd|th)-(\d+)(?:st|nd|rd|th)", combination, re.I)
    if m:
        return house_label(int(m.group(1)))
    m = re.search(r"Afflicted (\d+)(?:st|nd|rd|th) house", combination, re.I)
    if m:
        return house_label(int(m.group(1)))
    m = re.search(
        r"(\d+)(?:st|nd|rd|th) lord in (\d+)(?:st|nd|rd|th)",
        combination,
        re.I,
    )
    if m:
        return house_label(int(m.group(2)))
    m = re.search(r"(\d+)(?:st|nd|rd|th) lord", combination, re.I)
    if m:
        return house_label(int(m.group(1)))
    m = re.search(r"(\d+)(?:st|nd|rd|th)/(\d+)", combination)
    if m:
        return house_label(int(m.group(1)))
    if re.search(r"\bD-\d+\b|Dasha|Transit|Navamsa|Hora|Dashamsha", combination, re.I):
        return "General"
    if re.search(
        r"connected to|both afflicted|confirm|influence on|affecting|afflicted by",
        combination,
        re.I,
    ):
        nums = [int(x) for x in re.findall(r"(\d+)(?:st|nd|rd|th)", combination)]
        if nums:
            return house_label(nums[0])
    return "General"


def py_str(s: str) -> str:
    return repr(s)


def py_domains(domains: dict[str, dict[str, str]]) -> str:
    if not domains:
        return "{}"
    parts: list[str] = []
    for key in DOMAIN_KEYS:
        if key not in domains:
            continue
        d = domains[key]
        parts.append(
            f'        "{key}": {{'
            f'"area_affected": {py_str(d["area_affected"])}, '
            f'"impact": {py_str(d["impact"])}, '
            f'"severity": {py_str(d["severity"])}, '
            f'"conventional_remedies": {py_str(d["conventional_remedies"])}, '
            f'"modern_remedies": {py_str(d["modern_remedies"])}, '
            f"}},"
        )
    return "{\n" + "\n".join(parts) + "\n    }"


def main() -> None:
    house_matrix = load_house_severity_matrix()

    wb = openpyxl.load_workbook(XLSX_MAIN, read_only=True, data_only=True)
    sign_rows = read_sheet(wb, "Sign Wise Combos")
    house_rows = read_sheet(wb, "House Wise Combos")
    wb.close()

    # house_key -> row dict
    merged: dict[tuple[str, str], dict] = {}
    axis_index: dict[str, list[tuple[str, str]]] = {}

    def ensure_row(house: str, combination: str) -> dict:
        key = (house, combination)
        if key not in merged:
            merged[key] = {
                "house": house,
                "combination": combination,
                "domains": {},
            }
        canon = canonical_combination(combination)
        if is_axis_combination(combination):
            axis_index.setdefault(canon, [])
            if key not in axis_index[canon]:
                axis_index[canon].append(key)
        return merged[key]

    def attach_domain(
        row: dict,
        domain: str,
        domain_data: dict,
        remedy_combo: str,
    ) -> None:
        enrich_domain(domain_data, remedy_combo, domain, row["house"])
        row["domains"][domain] = domain_data

    for d in house_rows:
        house = d["House"]
        combo = d["Combination"]
        row = ensure_row(house, combo)
        health_data = {
            "area_affected": d["Health Focus"],
            "impact": d["Impact"],
            "severity": house_row_severity(house, combo, house_matrix),
        }
        attach_domain(row, "health", health_data, combo)

    domain_counts: dict[str, int] = {"health": len(house_rows)}

    for domain, (path, area_col, impact_col, sev_col) in DOMAIN_SHEETS.items():
        rows = read_domain_sheet(path)
        domain_counts[domain] = len(rows)
        for d in rows:
            combo_raw = d["Pitru Dosha Combination"]
            canon = canonical_combination(combo_raw)
            house = parse_house_label(combo_raw)
            domain_data = {
                "area_affected": d[area_col],
                "impact": d[impact_col],
                "severity": d[sev_col],
            }
            if is_axis_combination(combo_raw):
                targets = axis_index.get(canon)
                if targets:
                    for key in targets:
                        attach_domain(merged[key], domain, domain_data, combo_raw)
                    continue
            # Try merge with existing health row by canonical combo + house
            matched = False
            for key, row in merged.items():
                if key[0] == house and canonical_combination(row["combination"]) == canon:
                    attach_domain(row, domain, domain_data, combo_raw)
                    matched = True
                    break
            if not matched:
                row = ensure_row(house, combo_raw)
                attach_domain(row, domain, domain_data, combo_raw)

    house_wise = sorted(
        merged.values(),
        key=lambda r: (
            999 if r["house"] == "General" else int(re.match(r"^(\d+)", r["house"]).group(1)),
            r["combination"],
        ),
    )

    sign_wise = []
    for d in sign_rows:
        combo = d["Combination in Sign"]
        health_remedies = remedies_for_domain(combo, "health", "General", d["Health Impact"], d["Health Impact"])
        sign_wise.append(
            {
                "sign": d["Zodiac Sign"],
                "combination": combo,
                "stronger_houses": d["Stronger House/Axis"],
                "general_impact": d["Health Impact"],
                "nature_theme": d["Nature/Theme"],
                "severity": SIGN_COMBINATION_SEVERITY.get(combo, ""),
                "conventional_remedies": health_remedies["conventional_remedies"],
                "modern_remedies": health_remedies["modern_remedies"],
            }
        )

    lines = [
        '"""',
        "Pitru Dosha reference data (house-wise domains: health, career, finance, relationship).",
        "Sources: Pitru_Dosha_Zodiac_House_Wise_Combinations.xlsx + Career/Finance/Relationship workbooks.",
        "Regenerate: python backend/scripts/build_pitru_dosha_data.py",
        '"""',
        "",
        "from typing import TypedDict",
        "",
        "",
        "class DomainImpact(TypedDict):",
        "    area_affected: str",
        "    impact: str",
        "    severity: str",
        "    conventional_remedies: str",
        "    modern_remedies: str",
        "",
        "",
        "class SignWiseRow(TypedDict):",
        "    sign: str",
        "    combination: str",
        "    stronger_houses: str",
        "    general_impact: str",
        "    nature_theme: str",
        "    severity: str",
        "    conventional_remedies: str",
        "    modern_remedies: str",
        "",
        "",
        "class HouseWiseRow(TypedDict):",
        "    house: str",
        "    combination: str",
        "    domains: dict[str, DomainImpact]",
        "",
        "",
        "SIGN_WISE: list[SignWiseRow] = [",
    ]

    for row in sign_wise:
        lines.append(
            "    {"
            f'"sign": {py_str(row["sign"])}, '
            f'"combination": {py_str(row["combination"])}, '
            f'"stronger_houses": {py_str(row["stronger_houses"])}, '
            f'"general_impact": {py_str(row["general_impact"])}, '
            f'"nature_theme": {py_str(row["nature_theme"])}, '
            f'"severity": {py_str(row["severity"])}, '
            f'"conventional_remedies": {py_str(row["conventional_remedies"])}, '
            f'"modern_remedies": {py_str(row["modern_remedies"])}, '
            "},"
        )

    lines.append("]")
    lines.append("")
    lines.append("HOUSE_WISE: list[HouseWiseRow] = [")

    for row in house_wise:
        lines.append(
            "    {"
            f'"house": {py_str(row["house"])}, '
            f'"combination": {py_str(row["combination"])}, '
            f'"domains": {py_domains(row["domains"])}, '
            "},"
        )

    lines.append("]")
    lines.append("")
    lines.append("")
    lines.append("def sign_lookup(sign: str, combination: str) -> SignWiseRow | None:")
    lines.append("    for row in SIGN_WISE:")
    lines.append("        if row['sign'] == sign and row['combination'] == combination:")
    lines.append("            return row")
    lines.append("    return None")
    lines.append("")
    lines.append("")
    lines.append("def find_sign_row(")
    lines.append("    combination: str, signs: list[str]")
    lines.append(") -> tuple[SignWiseRow | None, str | None]:")
    lines.append("    for sign in signs:")
    lines.append("        row = sign_lookup(sign, combination)")
    lines.append("        if row:")
    lines.append("            return row, sign")
    lines.append("    return None, None")
    lines.append("")
    lines.append("")
    lines.append("def house_lookup(house_label: str, combination: str) -> HouseWiseRow | None:")
    lines.append("    for row in HOUSE_WISE:")
    lines.append("        if row['house'] == house_label and row['combination'] == combination:")
    lines.append("            return row")
    lines.append("    return None")
    lines.append("")
    lines.append("")
    lines.append("def combinations_for_sign(sign: str) -> list[str]:")
    lines.append("    return [r['combination'] for r in SIGN_WISE if r['sign'] == sign]")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    # Verify every domain row is stored
    stored_domain_rows = {k: 0 for k in DOMAIN_KEYS}
    for row in house_wise:
        for k in row["domains"]:
            stored_domain_rows[k] += 1

    missing_sign = {r["combination"] for r in sign_wise if not r["severity"]}
    print(f"Wrote {OUT} ({len(sign_wise)} sign, {len(house_wise)} house rows)")
    print("Domain rows loaded:", domain_counts)
    print("Domain rows stored (may exceed loaded when axis merges to multiple houses):", stored_domain_rows)
    if missing_sign:
        print("Sign combos without severity:", sorted(missing_sign))


if __name__ == "__main__":
    main()
